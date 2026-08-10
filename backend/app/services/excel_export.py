"""
Excel Export Service — Generate payroll Excel files matching the original template.

Produces:
  - Salary Detail sheet (薪资明细表)
  - Company Summary sheet (公司汇总表)
"""
import io
from datetime import datetime
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.models.all_models import SalaryRecord, Employee, Company


# ----- Styling constants -----
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="right", vertical="center")
TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CURRENCY_FORMAT = '#,##0.00'


def _style_header(ws, row: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _set_currency_cell(ws, row: int, col: int, value):
    """Set a cell with currency formatting."""
    cell = ws.cell(row=row, column=col, value=float(value) if value else 0)
    cell.number_format = CURRENCY_FORMAT
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER


def _set_text_cell(ws, row: int, col: int, value):
    """Set a text cell."""
    cell = ws.cell(row=row, column=col, value=str(value) if value else "")
    cell.alignment = TEXT_ALIGNMENT
    cell.border = THIN_BORDER


def generate_salary_detail(
    records: list[tuple[SalaryRecord, Employee, Optional[Company]]],
    period: str,
) -> io.BytesIO:
    """
    Generate the Salary Detail Excel sheet.

    Columns match the original Excel template structure.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "薪资明细"

    # Column headers — matching original Excel layout
    headers = [
        # Employee info
        "序号", "员工编号", "姓名", "公司全称", "部门", "岗位", "计税模式",
        # Income items
        "基本工资", "补贴", "考勤调整", "商保佣金", "KPI预提",
        "本月工资(F1)", "商办佣金", "绩效", "公寓佣金", "防暑降温费",
        "津贴", "保安奖金", "保洁奖金", "薪资小计(F2)",
        # Social insurance — personal
        "养老(个人)", "医疗(个人)", "失业(个人)", "公积金(个人)", "补充公积金(个人)",
        "个人福利合计(F3)",
        # Tax
        "累计应纳税所得额", "个税(F11/F14)",
        # Net pay
        "银行实发(F15)",
        # Social insurance — company
        "养老(公司)", "医疗(公司)", "失业(公司)", "工伤(公司)", "生育(公司)", "公积金(公司)",
        "公司福利合计(F17)",
        # Total cost
        "企业人力成本(F25)",
    ]

    header_col_count = len(headers)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=header_col_count)
    title_cell = ws.cell(row=1, column=1, value=f"开弈集团薪资明细表 — {period}")
    title_cell.font = Font(name="微软雅黑", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle (generation time)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=header_col_count)
    sub_cell = ws.cell(row=2, column=1,
                       value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    sub_cell.font = Font(name="微软雅黑", size=9)
    sub_cell.alignment = Alignment(horizontal="right")

    # Header row
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header(ws, 3, header_col_count)

    # Data rows
    total_row_count = 0
    for row_idx, (salary, emp, company) in enumerate(records, 4):
        col = 1
        _set_text_cell(ws, row_idx, col, row_idx - 3); col += 1
        _set_text_cell(ws, row_idx, col, emp.employee_no); col += 1
        _set_text_cell(ws, row_idx, col, emp.name); col += 1
        _set_text_cell(ws, row_idx, col, emp.company_full_name); col += 1
        _set_text_cell(ws, row_idx, col, emp.department or ""); col += 1
        _set_text_cell(ws, row_idx, col, emp.position or ""); col += 1
        _set_text_cell(ws, row_idx, col, emp.tax_type); col += 1

        _set_currency_cell(ws, row_idx, col, salary.base_salary); col += 1
        _set_currency_cell(ws, row_idx, col, salary.allowance); col += 1
        _set_currency_cell(ws, row_idx, col, salary.attendance_adjust); col += 1
        _set_currency_cell(ws, row_idx, col, salary.insurance_comm); col += 1
        _set_currency_cell(ws, row_idx, col, salary.kpi_provision); col += 1

        _set_currency_cell(ws, row_idx, col, salary.monthly_wage); col += 1
        _set_currency_cell(ws, row_idx, col, salary.office_comm); col += 1
        _set_currency_cell(ws, row_idx, col, salary.performance); col += 1
        _set_currency_cell(ws, row_idx, col, salary.apartment_comm); col += 1
        _set_currency_cell(ws, row_idx, col, salary.heat_allowance); col += 1

        _set_currency_cell(ws, row_idx, col, salary.other_allowance); col += 1
        _set_currency_cell(ws, row_idx, col, salary.security_bonus); col += 1
        _set_currency_cell(ws, row_idx, col, salary.cleaning_bonus); col += 1
        _set_currency_cell(ws, row_idx, col, salary.wage_subtotal); col += 1

        # Personal welfare — we store only total; individual items would need social_policy_detail
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 养老(个人) — placeholder
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 医疗(个人) — placeholder
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 失业(个人) — placeholder
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 公积金(个人) — placeholder
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 补充公积金 — placeholder

        _set_currency_cell(ws, row_idx, col, salary.personal_welfare); col += 1

        _set_currency_cell(ws, row_idx, col, salary.cumul_taxable_income); col += 1
        _set_currency_cell(ws, row_idx, col, salary.tax_amount); col += 1
        _set_currency_cell(ws, row_idx, col, salary.net_pay); col += 1

        # Company welfare — placeholder totals
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 养老(公司)
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 医疗(公司)
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 失业(公司)
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 工伤(公司)
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 生育(公司)
        _set_currency_cell(ws, row_idx, col, 0); col += 1  # 公积金(公司)

        _set_currency_cell(ws, row_idx, col, salary.company_welfare); col += 1
        _set_currency_cell(ws, row_idx, col, salary.total_cost); col += 1

        total_row_count = row_idx

    # Auto-adjust column widths
    for col in range(1, header_col_count + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Wider columns for text fields
    ws.column_dimensions["A"].width = 6   # 序号
    ws.column_dimensions["D"].width = 35  # 公司全称

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
